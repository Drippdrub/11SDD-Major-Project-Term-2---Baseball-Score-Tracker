# GUI Libraries
import tkinter as tk
from tkinter import messagebox as msgbox
import customtkinter as ctk
# Image Library
from PIL import Image
# Excel Read/Write Libraries
import xlsxwriter as xl
import openpyxl as openpx
# Excel To Image Converter
import excel2img
# OS
import os
import sys

if os.path.exists("./Game Results")==False: #creates 'Game Results' folder if it does not exist in the same directory as the program
    os.mkdir("./Game Results")

def resource_path(relative_path): #to access resource files as a --onefile
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# ctypes used solely for the purpose of importing custom fonts
try:
    from ctypes import windll, byref, create_unicode_buffer, create_string_buffer
    FR_PRIVATE  = 0x10
    FR_NOT_ENUM = 0x20
except ImportError: #for testing on mac, windll does not work on mac, disables the fonts on mac
    pass

global bgClr
bgClr = "#212121" #background colour used for the GUI

# this function loads in fonts from the /fontpath folder, taken from stackoverflow forums.

def loadfont(fontpath, private=True, enumerable=False):
    
    try: #for testing on mac, windll does not work on mac, disables the fonts on mac
        if isinstance(fontpath, bytes):
            pathbuf = create_string_buffer(fontpath)
            AddFontResourceEx = windll.gdi32.AddFontResourceExA
        elif isinstance(fontpath, str):
            pathbuf = create_unicode_buffer(fontpath)
            AddFontResourceEx = windll.gdi32.AddFontResourceExW
        else:
            raise TypeError('fontpath must be of type str or unicode')

        flags = (FR_PRIVATE if private else 0) | (FR_NOT_ENUM if not enumerable else 0)
        numFontsAdded = AddFontResourceEx(byref(pathbuf), flags, 0)
        return bool(numFontsAdded)
    except: #for testing on mac, windll does not work on mac, disables the fonts on mac
        pass

# root window class
class Win(ctk.CTk): 

    def __init__(self, *args, **kwargs):
        ctk.CTk.__init__(self, *args, **kwargs) #initialise window
        self.geometry("1280x720")
        self.resizable(False, False) #fixed window size
        self.iconbitmap(resource_path("ball_icon.ico")) #application icon
        self.title("Baseball/Softball Score Tracker") #window name
        #function allows custom fonts to be used
        loadfont(resource_path("Airstrike Academy.ttf"))
        loadfont(resource_path("Venus Plant.ttf"))
        loadfont(resource_path("Metropolis-Regular.otf"))
        #loading all fonts to be used in the GUI
        self.fontH1 = ctk.CTkFont(family="Airstrike Academy", size=72)
        self.fontH11 = ctk.CTkFont(family="Airstrike Academy", size=55)
        self.fontH2 = ctk.CTkFont(family="Venus Plant", size=32)
        self.fontB1 = ctk.CTkFont(family="Franklin Gothic", size=18)
        self.fontB2 = ctk.CTkFont(family="Cooper Black", size=24)
        self.fontB3 = ctk.CTkFont(family="Metropolis", size=15)
        self.fontB4 = ctk.CTkFont(family="Metropolis", size=13)

        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("dark-blue") #theme of window
        app = ctk.CTkFrame(self, fg_color=bgClr) #first frame, fills root window, parent of all other frames/screens
        app.pack(side="top", fill="both", expand=True)
        app.grid_rowconfigure(0, weight=1)
        app.grid_columnconfigure(0, weight=1)

        # packs all frames/screens to app frame
        self.frames = {}
        for F in (StartScreen, Lineups, Settings, GameScore, GameEnd):
            page_name = F.__name__
            frame = F(parent=app, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")
        
        # shows startscreen first (when __init__ runs)
        self.show_frame("StartScreen")
    
    # brings the parsed frame to the top (ie changes the displayed screen)
    def show_frame(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()
    

# main menu screen class
class StartScreen(ctk.CTkFrame):

    def __init__(self, parent, controller):
        ctk.CTkFrame.__init__(self, parent) #initialise frame
        self.controller = controller
        MMFrame = ctk.CTkFrame(self, fg_color=bgClr) #menu frame with Title and Menu Buttons (to the left of Main Menu Screen)
        # title
        MMTitle = ctk.CTkLabel(MMFrame, text="BASEBALL\nAND SOFTBALL\nSCORE\nTRACKER", font=controller.fontH1, bg_color=bgClr)
        MMTitle.pack(padx=30, pady=30)
        # start button
        MMStart = ctk.CTkButton(MMFrame, height=75, width=375, corner_radius=40, font=controller.fontH2,
                                text="Start Baseball Game", command=lambda: controller.show_frame("Lineups"))
        MMStart.pack(pady=50)
        # exit program button
        MMClose = ctk.CTkButton(MMFrame, height=75, width=375, corner_radius=40, font=controller.fontH2,
                                text="Close Program", command=lambda: self.destroyRequest())
        MMClose.pack(pady=50)

        MMFrame.pack(padx=50, side=tk.LEFT)

        # main menu image
        MMImage = ctk.CTkImage(dark_image=Image.open(resource_path("Baseball.png")), size=(400, 400))
        MMImgObj = ctk.CTkLabel(self, image=MMImage, text="")
        MMImgObj.pack(padx=50, side=tk.RIGHT)
    
    # asks user to confirm if they would like to leave
    def destroyRequest(self):
        close = msgbox.askyesno(title="Closing Program", message="Are you sure you want to exit the program?")
        if close == True:
            self.quit()

# Team Lineup Input Screen Class
class Lineups(ctk.CTkFrame):
    def __init__(self, parent, controller):
        ctk.CTkFrame.__init__(self, parent) #initialise frame
        self.controller = controller
        global HomeEntries #list of Home Team Player Entry Widgets, used to retrieve Player Names
        HomeEntries = []
        global AwayEntries #list of Away Team Player Entry Widgets, used to retrieve Player Names
        AwayEntries = []
        global Name1Ent #Home Team Name Widget, used to retrieve Home Team Name
        global Name2Ent #Away Team Name Widget, used to retrieve Away Team Name

        tLT1 = ctk.CTkFrame(self, fg_color=bgClr) #Frame, packs Home Team Entries to the left side of the screen
        tLT2 = ctk.CTkFrame(self, fg_color=bgClr) #Frame, packs Away Team Entries to the right side of the screen
        tLMid = ctk.CTkFrame(self, fg_color=bgClr) #Frame, separates the left and right sides of the screen, has change screen button

        tLT1.columnconfigure(0, weight=1) #Evenly weighted columns in .grid(), even distribution of cells
        tLT1.columnconfigure(1, weight=1)
        tLT1.columnconfigure(2, weight=1)

        T1Text = ctk.CTkLabel(tLT1, text="Home Team", font=controller.fontH2) #Left side, title for home team frame
        T1Text.grid(row=0, column=0, columnspan=3, sticky=tk.W+tk.E,
                    pady=10)
        T2Text = ctk.CTkLabel(tLT2, text="Away Team", font=controller.fontH2) #Right side, title for away team frame
        T2Text.grid(row=0, column=0, columnspan=3, sticky=tk.W+tk.E,
                    pady=10)


        Name1Lbl = ctk.CTkLabel(tLT1, text="Team Name:", font=controller.fontB2) #Left side, label for home team name entry
        Name1Lbl.grid(row=1, column=0, sticky=tk.E,
                    padx=10, pady=10)
        Name1Ent = ctk.CTkEntry(tLT1, placeholder_text="Home Team", font=controller.fontB1, height=40) #Left side, entry for home team name
        Name1Ent.grid(row=1, column=1, columnspan=2, sticky=tk.W+tk.E,
                    padx=40, pady=10)
        for i in range(9): #nine players on home team
            PlayerName = ctk.CTkLabel(tLT1, text=f"Player {i+1} Name:", font=controller.fontB2) #Left side, label for player name entries
            PlayerName.grid(row=i+2, column=0, sticky=tk.E,
                            padx=10, pady=10)
            P1NameEnt = ctk.CTkEntry(tLT1, placeholder_text=f"Player {i+1}", font=controller.fontB1, height=40) #Left side, entry for player names
            P1NameEnt.grid(row=i+2, column=1, columnspan=2, sticky=tk.W+tk.E,
                        padx=40, pady=10)
            HomeEntries.append(P1NameEnt) #Add player name entry widgets to a list to grab values

        tLT2.columnconfigure(0, weight=1) #Evenly weighted columns in .grid(), even distribution of cells
        tLT2.columnconfigure(1, weight=1)
        tLT2.columnconfigure(2, weight=1)

        Name2Lbl = ctk.CTkLabel(tLT2, text="Team Name:", font=controller.fontB2) #Right side, label for away team name entry
        Name2Lbl.grid(row=1, column=0, sticky=tk.E,
                    padx=10, pady=10)
        Name2Ent = ctk.CTkEntry(tLT2, placeholder_text="Away Team", font=controller.fontB1, height=40) #Right side, entry for away team name
        Name2Ent.grid(row=1, column=1, columnspan=2, sticky=tk.W+tk.E,
                    padx=40, pady=10)
        for i in range(9): #nine players on away team
            PlayerName = ctk.CTkLabel(tLT2, text=f"Player {i+1} Name:", font=controller.fontB2) #Right side, label for player name entries
            PlayerName.grid(row=i+2, column=0, sticky=tk.E,
                            padx=10, pady=10)
            P2NameEnt = ctk.CTkEntry(tLT2, placeholder_text=f"Player {i+1}", font=controller.fontB1, height=40) #Right side, entry for player names
            P2NameEnt.grid(row=i+2, column=1, columnspan=2, sticky=tk.W+tk.E,
                        padx=40, pady=10)
            AwayEntries.append(P2NameEnt) #Add player name entry widgets to a list to grab values



        tLT1.pack(padx=10, side=tk.LEFT, fill=tk.BOTH, expand=True) #pack home team entries to left side
        tLT2.pack(padx=10, side=tk.RIGHT, fill=tk.BOTH, expand=True) #pack away team entries to right side

        tLText = ctk.CTkLabel(tLMid, text="Make sure the players have been ordered into batting order",
                                font=controller.fontB1, wraplength=300, justify=tk.CENTER) #Middle of screen, label to remind users to order players in batting order
        tLStartGame = ctk.CTkButton(tLMid, text="Start", command=lambda: self.proceedRequest(controller)) #Middle of screen, button to move to settings

        tLText.pack(pady=15, anchor=tk.CENTER) #pack to center
        tLStartGame.pack(pady=30, anchor=tk.CENTER) #pack to center
        self.update()
        tLMid.pack(pady=700/2-200, anchor=tk.CENTER) #Middle frame is in the center of the screen

    def proceedRequest(self, controller):
        self.controller = controller #controller is the controller of __init__ function, Win() class
        #confirm with user that input choices are correct
        proceed = msgbox.askokcancel(title="Confirm Action", message="Do you wish to proceed?\n\nTeam batting orders cannot be changed from this point onward.")
        #if user confirms
        if proceed:
            global HomeMembers
            HomeMembers = []
            for entry in HomeEntries:
                HomeMembers.append(entry.get()) #grabs values from home team player name entries, and adds to list
            global AwayMembers
            AwayMembers = []
            for entry in AwayEntries:
                AwayMembers.append(entry.get()) #grabs values from away team player name entries, and adds to list
            global HomeName
            HomeName = Name1Ent.get() #grabs home team name
            global AwayName
            AwayName = Name2Ent.get() #grabs home team name
            if HomeName=="" or AwayName=="":
                #check if team names have been entered
                msgbox.showerror(title="Error",
                                     message="The Team Name entry fields are empty.\nPlease check that all fields have been filled in.")
                return
            for member in HomeMembers:
                if member=="":
                    #check if home team player names have been entered
                    msgbox.showerror(title="Error",
                                     message="One or more entry fields under Team 1 are empty.\nPlease check that all fields have been filled in.")
                    return
            for member in AwayMembers:
                if member=="":
                    #check if away team player names have been entered
                    msgbox.showerror(title="Error",
                                     message="One or more entry fields under Team 2 are empty.\nPlease check that all fields have been filled in.")
                    return
            #swap frame to Settings() class
            controller.show_frame("Settings")

#Settings frame class
class Settings(ctk.CTkFrame):
    
    def __init__(self, parent, controller):
        ctk.CTkFrame.__init__(self, parent) #initialise frame
        #second frame, with padding
        settingFrame = ctk.CTkFrame(self)

        #variable for filename of excel spreadsheet
        fileName = tk.StringVar()

        fileNameLbl = ctk.CTkLabel(settingFrame, text="Excel File Name:") #label from file name entry
        fileNameEnt = ctk.CTkEntry(settingFrame, textvariable=fileName) #entry for file name var
        fileNameLbl.grid(row=0, column=0, padx=10, pady=10)
        fileNameEnt.grid(row=0, column=1, padx=10, pady=10)
        #label to tell users about exported excel spreadsheet
        fileInfoLbl = ctk.CTkLabel(self, text="Game results will be automatically\noutput to an excel worksheet in\nthe 'Game Results' folder.")
        fileInfoLbl.place(relx=0.75, rely=0.45)

        gameLengthLbl = ctk.CTkLabel(settingFrame, text="Select A Game Length:") #label for game length buttons
        gameLengthLbl.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

        #variable for length (in innings) of the baseball/softball game
        global gameLength
        gameLength = tk.IntVar(value=0)
        Int7Btn = ctk.CTkRadioButton(settingFrame, text="Softball Game (7 Innings)", variable= gameLength, value=7) #Radio button for 7 innings
        Int9Btn = ctk.CTkRadioButton(settingFrame, text="Baseball Game (9 Innings)", variable= gameLength, value=9) #Radio button for 9 innings
        Int7Btn.grid(row=2, column=0, padx=10, pady=10)
        Int9Btn.grid(row=2, column=1, padx=10, pady=10)

        #button to proceed to next screen (GameScore() frame)
        settingButton = ctk.CTkButton(settingFrame, text="Proceed", command=lambda: self.openGamescore(controller, fileNameEnt.get()))
        settingButton.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
        settingFrame.pack(expand=True, anchor=tk.CENTER)
    
    def openGamescore(self, controller, fileName):
        if fileName=="":
            #check if a file name has been given
            msgbox.showerror(title="Error",
                                     message="An excel filename has not been selected.\nPlease check you have given a filename.")
            return
        if gameLength.get()==0:
            #check if a gameLength is selected, default value of gameLength is 0
            return

        #xlfile - path of excel spreadsheet
        global xlfile
        #fileTitle - fileName of excel spreadsheet
        global fileTitle
        fileTitle = fileName
        xlfile = f"Game Results/{fileName}.xlsx"

        #create the file with xlsxwriter
        bsbTrack = xl.Workbook(xlfile)
        playerScores = bsbTrack.add_worksheet()
        #close the file with xlsxwriter
        bsbTrack.close()

        #table - excel workbook
        global table
        #sheet - excel worksheet inside workbook
        global sheet

        table = openpx.load_workbook(xlfile)
        sheet = table.active
        #rename excel worksheet to 'playerScores'
        sheet.title = "playerScores"

        #change column width of columns in excel spreadsheet
        for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            sheet.column_dimensions[column].width = 15

        #template for excel spreadsheet layout
        tableTemplate = [
            ["Home Team", HomeName, "", "", "", "", "", "Away Team", AwayName, "", "", "", ""],
            ["Player ID", "Player Name", "Runs", "Strikes", "Foul Balls", "Balls", "", "Player ID", "Player Name", "Runs", "Strikes", "Foul Balls", "Balls"],
            [1, "", 0, 0, 0, 0, "", 1, "", 0, 0, 0, 0],
            [2, "", 0, 0, 0, 0, "", 2, "", 0, 0, 0, 0],
            [3, "", 0, 0, 0, 0, "", 3, "", 0, 0, 0, 0],
            [4, "", 0, 0, 0, 0, "", 4, "", 0, 0, 0, 0],
            [5, "", 0, 0, 0, 0, "", 5, "", 0, 0, 0, 0],
            [6, "", 0, 0, 0, 0, "", 6, "", 0, 0, 0, 0],
            [7, "", 0, 0, 0, 0, "", 7, "", 0, 0, 0, 0],
            [8, "", 0, 0, 0, 0, "", 8, "", 0, 0, 0, 0],
            [9, "", 0, 0, 0, 0, "", 9, "", 0, 0, 0, 0],
            ["Team Total", "", "", "", "", "", "", "Team Total", "", "", "", "", ""],
        ]

        #add template to excel spreadsheet
        for row in tableTemplate:
            sheet.append(row)

        #add home team player names to excel spreadsheet
        for player in HomeMembers:
            index = HomeMembers.index(player)
            cell = f"B{index+3}"
            sheet[cell] = player
        
        #add away team player names to excel spreadsheet
        for player in AwayMembers:
            index = AwayMembers.index(player)
            cell = f"I{index+3}"
            sheet[cell] = player

        #save excel spreadsheet
        table.save(xlfile)

        #switch screen to GameScore() class
        controller.show_frame("GameScore")

        #update values in GameScore() class
        global homeTeamName
        homeTeamName = HomeName #different variable, same purpose
        global awayTeamName
        awayTeamName = AwayName #different variable, same purpose

        #update widgets in GameScore() class
        #send home and away team names to BattingTeam widget, set value to away team
        BattingTeam.configure(values=[f"Home Team ({homeTeamName})", f"Away Team ({awayTeamName})"], width=200)
        BattingTeam.set(f"Away Team ({awayTeamName})")
        BattingTeam.configure(state="disabled") #setting value of comboBox while disabled will not work
        #display away team as batting first, home team as fielding first
        batText.configure(text=f"Away Team ({awayTeamName})")
        fldText.configure(text=f"Home Team ({homeTeamName})")
        #list of batters updated
        batterEntry.configure(values=AwayMembers)
        batterEntry.set(AwayMembers[0])
        #list of pitchers updated
        fieldEntry.configure(values=HomeMembers)
        fieldEntry.set(HomeMembers[0])



class GameScore(ctk.CTkFrame):

    def __init__(self, parent, controller):
        #number of times teams have swapped sides, value is even when away team is batting, odd when home team is batting
        global teamSwapNum
        teamSwapNum = 0

        #set home and away team names to default values (this is run on program start/initialisation, set to proper value later)
        global homeTeamName
        homeTeamName = ""
        global awayTeamName
        awayTeamName = ""

        #globalising certain widgets to be set and grabbed by other functions
        global BattingTeam
        global batText
        global fldText

        #inningNum - current inning of game
        global inningNum
        inningNum = tk.StringVar()
        inningNum.set(1)
        #trace values of innings in ent, change value to old value if a non-number character is entered
        inningNum.trace("w", lambda name, index, mode, inningNum=inningNum: self.innCallback(inningNum, inningOld))
        inningOld = inningNum
        
        ctk.CTkFrame.__init__(self, parent) #initialise frame
        self.controller = controller
        upArrow = ctk.CTkImage(dark_image=Image.open(resource_path("up_arrow.png")), size=(10, 10)) #image of up arrow
        downArrow = ctk.CTkImage(dark_image=Image.open(resource_path("down_arrow.png")), size=(10, 10)) #image of down arrow

        topFrame = ctk.CTkFrame(master=self) #frame at the top of GameScore screen
        topFrame.pack(padx=20, pady=10, fill=tk.X)

        #label of inning entry
        inningLbl = ctk.CTkLabel(master=topFrame, text="Inning:", font=controller.fontB1)
        inningLbl.grid(row=0, column=0, rowspan=2, padx=5)
        #entry of inningNum var
        inningEnt = ctk.CTkEntry(master=topFrame, font=controller.fontB1, height=40, textvariable=inningNum)
        inningEnt.grid(row=0, column=1, rowspan=2)
        #button to increase inningNum by 1
        inningUp = ctk.CTkButton(master=topFrame, image=upArrow, text="", 
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entAdd(inningNum, controller, "inningNum"))
        #button to decrease inningNum by 1
        inningDown = ctk.CTkButton(master=topFrame, image=downArrow, text="",
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entSub(inningNum, "inningNum"))
        inningUp.grid(row=0, column=2, padx=4, pady=2)
        inningDown.grid(row=1, column=2, pady=2)

        #label for current batting team
        BattingTeamLbl = ctk.CTkLabel(topFrame, text="Current Batting Team: ")
        BattingTeamLbl.grid(row=0, column=3, rowspan=2, padx=15)

        #dropdown menu for current battingteam, dropdown function is disabled but was used during testing
        BattingTeam = ctk.CTkComboBox(topFrame, values=[f"Home Team ({homeTeamName})", f"Away Team ({awayTeamName})"],
                                      text_color_disabled="gray84")
        #set away team to bat first
        BattingTeam.set(f"Away Team ({awayTeamName})")
        BattingTeam.grid(row=0, column=4, rowspan=2, padx=5)
        #add command for button press
        BattingTeam.configure(command=lambda e: self.updateTeams(batterEntry, fieldEntry))

        #re-enables batting team dropdown menu for testing purposes
        devModeOn = ctk.IntVar(value=0)
        devOptions = ctk.CTkCheckBox(topFrame, text="Enable DevMode?", command=lambda: self.updateDevMode(devOptions, BattingTeam), variable=devModeOn, onvalue=1, offvalue=0)
        #disabled packing this checkbox
        # devOptions.grid(row=0, column=7, rowspan=2, padx=15)

        tabview = ctk.CTkTabview(master=self, command=lambda: self.updateTable())
        tabview.pack(padx=20, pady=5, expand=True, fill=tk.BOTH)

        allTab = tabview.add("Game Scores")
        overviewTab = tabview.add("Game Overview")
        tabview.set("Game Scores")

        allTab.grid_columnconfigure(0, weight=1)
        allTab.grid_columnconfigure(1, weight=1)
        allTab.grid_rowconfigure(0, weight=1)

        #All Tab
        AllBat = ctk.CTkFrame(allTab, fg_color=bgClr)
        AllBat.columnconfigure(0, weight=2)
        AllBat.columnconfigure(1, weight=3)
        AllBat.columnconfigure(2, weight=1)
        AllBat.columnconfigure(3, weight=2)
        AllBat.columnconfigure(4, weight=3)
        AllBat.columnconfigure(5, weight=1)
        AllBat.columnconfigure(6, weight=2)
        AllBat.columnconfigure(7, weight=2)
        AllBat.columnconfigure(8, weight=2)
        AllFld = ctk.CTkFrame(allTab, fg_color=bgClr)
        AllFld.columnconfigure(0, weight=2)
        AllFld.columnconfigure(1, weight=3)
        AllFld.columnconfigure(2, weight=1)
        AllFld.columnconfigure(3, weight=2)
        AllFld.columnconfigure(4, weight=3)
        AllFld.columnconfigure(5, weight=1)
        AllFld.columnconfigure(6, weight=2)
        AllFld.columnconfigure(7, weight=2)
        AllFld.columnconfigure(8, weight=2)

        batText = ctk.CTkLabel(AllBat, text=BattingTeam.get(), font=controller.fontB1)
        batText.grid(row=0, column=0, columnspan=9, sticky=tk.W+tk.E,
                    pady=20)
        fldText = ctk.CTkLabel(AllFld, text=f"Fielding Team ({homeTeamName})", font=controller.fontB1)
        fldText.grid(row=0, column=0, columnspan=9, sticky=tk.W+tk.E,
                    pady=20)

        #All Tab Batting Frame
        global batterEntry

        batterName = ctk.CTkLabel(AllBat, text="Batter:", font=controller.fontB3)
        batterEntry = ctk.CTkComboBox(AllBat)
        batterName.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)
        batterEntry.grid(row=1, column=3, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)
        blankLbl = ctk.CTkLabel(AllBat, text="", width=100)
        blankLbl.grid(row=1, column=6, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)

        global runs
        runs = tk.StringVar()
        runs.set(0)
        runs.trace("w", lambda name, index, mode, runs=runs: self.callback(runs, runsOld))
        runsOld = ""

        runText = ctk.CTkLabel(AllBat, text="Runs:", font=controller.fontB3)
        runEntry = ctk.CTkEntry(AllBat, textvariable=runs, width=50)
        runText.grid(row=2, rowspan=2, column=0, sticky=tk.E, padx=10, pady=10)
        runEntry.grid(row=2, rowspan=2, column=1, sticky=tk.W+tk.E, padx=5, pady=10)
        runUp = ctk.CTkButton(master=AllBat, image=upArrow, text="", 
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entAdd(runs, controller))
        runDown = ctk.CTkButton(master=AllBat, image=downArrow, text="",
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entSub(runs))
        runUp.grid(row=2, column=2, padx=4, pady=2, sticky=tk.W)
        runDown.grid(row=3, column=2, padx=4, pady=2, sticky=tk.W)

        global strikes
        strikes = tk.StringVar()
        strikes.set(0)
        strikes.trace("w", lambda name, index, mode, strikes=strikes: self.callback(strikes, strikesOld))
        strikesOld = ""

        strikeText = ctk.CTkLabel(AllBat, text="Strikes:", font=controller.fontB3)
        strikeEntry = ctk.CTkEntry(AllBat, textvariable=strikes, width=50)
        strikeText.grid(row=2, rowspan=2, column=3, sticky=tk.E, padx=10, pady=10)
        strikeEntry.grid(row=2, rowspan=2, column=4, sticky=tk.W+tk.E, padx=5, pady=10)
        strikeUp = ctk.CTkButton(master=AllBat, image=upArrow, text="", 
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entAdd(strikes, controller, "strikes", foulBall))
        strikeDown = ctk.CTkButton(master=AllBat, image=downArrow, text="",
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entSub(strikes, "strikes"))
        strikeUp.grid(row=2, column=5, padx=4, pady=2, sticky=tk.W)
        strikeDown.grid(row=3, column=5, padx=4, pady=2, sticky=tk.W)

        global foulBall
        foulBall = tk.StringVar()
        foulBall.set(0)
        foulBall.trace("w", lambda name, index, mode, foulBall=foulBall: self.callback(foulBall, foulBallOld))
        foulBallOld = ""

        foulBallText = ctk.CTkLabel(AllBat, text="Foul Balls:", font=controller.fontB3)
        foulBallEntry = ctk.CTkEntry(AllBat, textvariable=foulBall, width=50)
        foulBallText.grid(row=4, rowspan=2, column=0, sticky=tk.E, padx=10, pady=30)
        foulBallEntry.grid(row=4, rowspan=2, column=1, sticky=tk.W+tk.E, padx=5, pady=10)
        foulBallUp = ctk.CTkButton(master=AllBat, image=upArrow, text="", 
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entAdd(foulBall, controller, "foulBall", strikes))
        foulBallDown = ctk.CTkButton(master=AllBat, image=downArrow, text="",
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entSub(foulBall, "foulBall"))
        foulBallUp.grid(row=4, column=2, padx=4, pady=3, sticky=tk.W+tk.S)
        foulBallDown.grid(row=5, column=2, padx=4, pady=3, sticky=tk.W+tk.N)

        global battersOut
        battersOut = 0
        global batOutText
        batOutText = ctk.CTkLabel(AllBat, text=f"Batters Out: {battersOut}", font=controller.fontB3)
        batOutText.grid(row=4, rowspan=2, column=3, columnspan=2, sticky=tk.E, padx=10, pady=10)
        

        # All Tab Fielding Frame
        global fieldEntry

        fieldName = ctk.CTkLabel(AllFld, text="Pitcher:", font=controller.fontB3)
        fieldEntry = ctk.CTkComboBox(AllFld)
        fieldName.grid(row=1, column=0, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)
        fieldEntry.grid(row=1, column=3, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)
        blankLbl2 = ctk.CTkLabel(AllFld, text="", width=130)
        blankLbl2.grid(row=1, column=6, columnspan=3, sticky=tk.W+tk.E, padx=20, pady=30)

        global balls
        balls = tk.StringVar()
        balls.set(0)
        balls.trace("w", lambda name, index, mode, balls=balls: self.callback(balls, ballsOld))
        ballsOld = ""

        ballText = ctk.CTkLabel(AllFld, text="Balls:", font=controller.fontB3)
        ballEntry = ctk.CTkEntry(AllFld, textvariable=balls, width=50)
        ballText.grid(row=2, rowspan=2, column=0, sticky=tk.E, padx=10, pady=10)
        ballEntry.grid(row=2, rowspan=2, column=1, sticky=tk.W+tk.E, padx=5, pady=10)
        ballUp = ctk.CTkButton(master=AllFld, image=upArrow, text="", 
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entAdd(balls, controller, "balls"))
        ballDown = ctk.CTkButton(master=AllFld, image=downArrow, text="",
                                width=10, height=10, fg_color="#ababab",
                                command=lambda: self.entSub(balls))
        ballUp.grid(row=2, column=2, padx=4, pady=2, sticky=tk.W)
        ballDown.grid(row=3, column=2, padx=4, pady=2, sticky=tk.W)

        AllBat.grid(row=0, column=0, sticky="nsew", padx=5)
        AllFld.grid(row=0, column=1, sticky="nsew", padx=5)
        # AllBat.pack(padx=(10, 5), side=tk.LEFT, fill=tk.BOTH, expand=True)
        # AllFld.pack(padx=(5, 10), side=tk.RIGHT, fill=tk.BOTH, expand=True)

        addStatsBtn = ctk.CTkButton(allTab, text="Add Scores to Player", command=lambda: self.changeBatter(controller))
        addStatsBtn.place(relx=0.5, rely=0.8, anchor=tk.CENTER)

        earlyStopBtn = ctk.CTkButton(allTab, text="End Game Early", fg_color="#f02d13",command=lambda: self.earlyGameEnd(controller))
        earlyStopBtn.place(relx=0.8, rely=0.8, anchor=tk.CENTER)

        #overview Tab
        global overviewFrame
        overviewFrame = ctk.CTkFrame(overviewTab)
        overviewFrame.pack(fill="both", expand=True)

        global overviewTableValues
        overviewTableValues = []
        for x in range(12):
            overviewFrame.rowconfigure(x, weight=1)
            for y in range(13):
                overviewFrame.columnconfigure(y, weight=1)
                if (y==1 and x==0) or (y==8 and x==0):
                    e=ctk.CTkEntry(overviewFrame,
                                font=controller.fontB4, corner_radius=0)
                    e.grid(row=x, column=y, sticky="nsew", columnspan=2)
                elif (y==2 and x==0) or (y==9 and x==0):
                    pass
                else:
                    e=ctk.CTkEntry(overviewFrame,
                                font=controller.fontB4, corner_radius=0)
                    e.grid(row=x, column=y, sticky="nsew")
                overviewTableValues.append(e)
                
        
    
    def innCallback(self, sv, old):
        if sv.get().isdigit():
            old = sv.get()
        else:
            sv.set(old)
    
    def callback(self, sv, old):
        if sv.get().isdigit():
            old = sv.get()
        else:
            sv.set(old)

    def entAdd(self, var, controller, *args):
        global battersOut
        try:
            name = args[0]
        except:
            name = "none"
        try:
            var2 = int(args[1].get())
        except:
            var2 = 0
        if name == "inningNum":
            if int(var.get()) < gameLength.get():
                var.set(int(var.get())+1)
                global teamSwapNum
                if BattingTeam.get()==f"Away Team ({awayTeamName})":
                    teamSwapNum = (2*int(var.get())) - 2
                else:
                    teamSwapNum = (2*int(var.get())) - 1

            elif int(var.get()) >= gameLength.get():
                endGame = msgbox.askyesno(title="Finishing Game", message="Do you wish to proceed?\n\nThis will end the game, or move the game into overtime.")
                if endGame:
                    if sheet["C12"].value == sheet["J12"].value:
                        var.set(int(var.get())+1)
                        if int(var.get()) == 8:
                            msgbox.showinfo(title="Overtime", message="Game has entered overtime.")
                        else:
                            msgbox.showinfo(title="Overtime", message="Overtime has continued.")
                    else:
                        controller.show_frame("GameEnd")
                        GameEnd.updateWinScreen(GameEnd)
        elif name == "strikes":
            if int(var.get())+var2 < 2:
                var.set(int(var.get())+1)
            elif int(var.get())+var2 >= 2:
                strikeOut = msgbox.askokcancel(title="Batter out", message="The current batter will become out.\nPress cancel to undo.")
                if strikeOut:
                    var.set(int(var.get())+1)
                    battersOut += 1
                    batOutText.configure(text=f"Batters Out: {battersOut}")
                    self.changeBatter(controller)
        elif name == "foulBall":
            if var2==2:
                foulOut = msgbox.askokcancel(title="Batter out", message="The current batter will become out.\nPress cancel to undo.")
                if foulOut:
                    var.set(int(var.get())+1)
                    battersOut += 1
                    batOutText.configure(text=f"Batters Out: {battersOut}")
                    self.changeBatter(controller)
            elif var2==1 and int(var.get())==1:
                foulOut = msgbox.askokcancel(title="Batter out", message="The current batter will become out.\nPress cancel to undo.")
                if foulOut:
                    var.set(int(var.get())+1)
                    battersOut += 1
                    batOutText.configure(text=f"Batters Out: {battersOut}")
                    self.changeBatter(controller)
            else:
                var.set(int(var.get())+1)
        elif name == "balls":
            if int(var.get())+var2 < 4:
                var.set(int(var.get())+1)
            elif int(var.get())+var2 >= 4:
                strikeOut = msgbox.askokcancel(title="Batter walk", message="The current batter will walk to next base\nPress cancel to undo.")
                if strikeOut:
                    var.set(int(var.get())+1)
                    self.changeBatter(controller)
        else:
            var.set(int(var.get())+1)

    def entSub(self, var, *args):
        try:
            name = args[0]
        except:
            name = "none"
        if name == "inningNum":
            if int(var.get()) > 1:
                var.set(int(var.get())-1)
        else:
            if int(var.get()) > 0:
                var.set(int(var.get())-1)

    def changeBatter(self, controller):
    
        index = batterEntry.cget("values").index(batterEntry.get())
        if BattingTeam.get()==f"Away Team ({awayTeamName})":
            for row in range(0, 9):
                if index==row:
                    sheet[f"J{row+3}"] = int(runs.get())
                    sheet[f"K{row+3}"] = int(strikes.get())
                    sheet[f"L{row+3}"] = int(foulBall.get())
                if fieldEntry.cget("values").index(fieldEntry.get())==row:
                    sheet[f"F{row+3}"] = int(balls.get())
        else:
            for row in range(0, 9):
                if index==row:
                    sheet[f"C{row+3}"] = int(runs.get())
                    sheet[f"D{row+3}"] = int(strikes.get())
                    sheet[f"E{row+3}"] = int(foulBall.get())
                if fieldEntry.cget("values").index(fieldEntry.get())==row:
                    sheet[f"M{row+3}"] = int(balls.get())
        for column in ["C", "D", "E", "F", "J", "K", "L", "M"]:
            colTotal = 0
            for i in range(3, 11):
                colTotal += int(sheet[f"{column}{i}"].value)
            sheet[f"{column}12"] = colTotal
        runs.set(0)
        strikes.set(0)
        foulBall.set(0)
        balls.set(0)

        table.save(xlfile)

        try:
            batterEntry.set(batterEntry.cget("values")[index+1])
        except IndexError:
            batterEntry.set(batterEntry.cget("values")[0])

        global battersOut
        if battersOut >= 3:
            msgbox.showinfo(title="Swapping Sides", message="Teams will now swap sides.")
            battersOut = 0
            batOutText.configure(text=f"Batters Out: {battersOut}")
            self.swapSides(controller)
    
    def swapSides(self, controller):
        BattingTeam.configure(state="normal")
        if BattingTeam.get()==f"Away Team ({awayTeamName})":
            BattingTeam.set(f"Home Team ({homeTeamName})")
        else:
            BattingTeam.set(f"Away Team ({awayTeamName})")
        BattingTeam.configure(state="disabled")
        global teamSwapNum
        teamSwapNum += 1
        if teamSwapNum%2==0:
            self.entAdd(inningNum, controller, "inningNum", None)
        self.updateTeams(batterEntry, fieldEntry)
    
    def earlyGameEnd(self, controller):
        endConfirm = msgbox.askyesno(title="End Game Early?", message="Are you sure you want to end the game early?")
        if endConfirm:
            index = batterEntry.cget("values").index(batterEntry.get())
            if BattingTeam.get()==f"Away Team ({awayTeamName})":
                for row in range(0, 8):
                    if index==row:
                        sheet[f"J{row+3}"] = int(runs.get())
                        sheet[f"K{row+3}"] = int(strikes.get())
                        sheet[f"L{row+3}"] = int(foulBall.get())
                    if fieldEntry.cget("values").index(fieldEntry.get())==row:
                        sheet[f"F{row+3}"] = int(balls.get())
            else:
                for row in range(0, 8):
                    if index==row:
                        sheet[f"C{row+3}"] = int(runs.get())
                        sheet[f"D{row+3}"] = int(strikes.get())
                        sheet[f"E{row+3}"] = int(foulBall.get())
                    if fieldEntry.cget("values").index(fieldEntry.get())==row:
                        sheet[f"M{row+3}"] = int(balls.get())
            for column in ["C", "D", "E", "F", "J", "K", "L", "M"]:
                colTotal = 0
                for i in range(3, 11):
                    colTotal += int(sheet[f"{column}{i}"].value)
                sheet[f"{column}12"] = colTotal

            table.save(xlfile)

            controller.show_frame("GameEnd")
            GameEnd.updateWinScreen(GameEnd)

    def updateDevMode(self, devOptions, BattingTeam):
        if devOptions.get() == 1:
            BattingTeam.configure(state="normal")
        else:
            BattingTeam.configure(state="disabled")
    
    def updateTable(self):
        overviewList = []
        overviewTable = list(sheet.values)
        for row in overviewTable:
            for term in row:
                overviewList.append(term)

        for cell in overviewTableValues:
            index = overviewTableValues.index(cell)
            cell.configure(state="normal")
            cell.delete(0, tk.END)
            cell.insert(1, str(overviewList[index]))
            cell.configure(state="disabled")

    def updateTeams(self, batterName, fieldName):
        batOld=batText.cget("text")
        text=BattingTeam.get()
        batText.configure(text=text)
        fldText.configure(text=batOld)
        if text == f"Home Team ({homeTeamName})":
            batterName.configure(values=HomeMembers)
            batterName.set(HomeMembers[0])
            fieldName.configure(values=AwayMembers)
            fieldName.set(AwayMembers[0])
        else:
            batterName.configure(values=AwayMembers)
            batterName.set(AwayMembers[0])
            fieldName.configure(values=HomeMembers)
            fieldName.set(HomeMembers[0])

class GameEnd(ctk.CTkFrame):
    
    def __init__(self, parent, controller):
        ctk.CTkFrame.__init__(self, parent)
        gameEndFrame = ctk.CTkFrame(self)
        gameEndFrame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        gameEndFrame.grid_rowconfigure(0, weight=1)
        gameEndFrame.grid_rowconfigure(1, weight=1)
        gameEndFrame.grid_rowconfigure(2, weight=4)
        gameEndFrame.grid_columnconfigure(0, weight=1)
        gameEndFrame.grid_columnconfigure(1, weight=1)

        global endTitle
        endTitle = ctk.CTkLabel(gameEndFrame, text="", font=controller.fontH1)
        endTitle.grid(row=0, column=0, columnspan=2, pady=15)

        exportImgBtn = ctk.CTkButton(gameEndFrame, text="Export Game Results As Image", command=self.exportAsImg)
        exportImgBtn.grid(row=1, column=0, padx=5, pady=15, sticky=tk.E)
        exportImgInfo = ctk.CTkLabel(gameEndFrame, text="The excel table of game scores\nwill be exported as an image\nin the 'Game Results' folder.")
        exportImgInfo.grid(row=1, column=1, padx=5, pady=15, sticky=tk.W)

        resultsFrame = ctk.CTkFrame(gameEndFrame)
        resultsFrame.grid(row=2, column=0, columnspan=2)

        global resultsValues
        resultsValues = []
        for x in range(14):
            resultsFrame.rowconfigure(x, weight=1)
            for y in range(13):
                resultsFrame.columnconfigure(y, weight=1)
                if (y==1 and x==2) or (y==8 and x==2) or (y==1 and x==0):
                    e=ctk.CTkEntry(resultsFrame,
                                font=controller.fontB4, corner_radius=0)
                    e.grid(row=x, column=y, sticky="nsew", columnspan=2)
                elif (y==2 and x==2) or (y==9 and x==2) or (y==2 and x==0):
                    pass
                else:
                    e=ctk.CTkEntry(resultsFrame,
                                font=controller.fontB4, corner_radius=0)
                    e.grid(row=x, column=y, sticky="nsew")
                resultsValues.append(e)
        
        restartBtn = ctk.CTkButton(gameEndFrame, text="Back to Main Menu", command=lambda: controller.show_frame("StartScreen"))
        restartBtn.grid(row=3, column=0, padx=5, pady=15, sticky=tk.E)
        closeBtn = ctk.CTkButton(gameEndFrame, text="Close Program", command=self.destroyRequest)
        closeBtn.grid(row=3, column=1, padx=5, pady=15, sticky=tk.W)
    
    def exportAsImg(self):  
        try:
            excel2img.export_img(xlfile,f"Game Results/{fileTitle}.png", "", "playerScores!A1:M14")
            msgbox.showinfo(title="Image Successfully Created", message="Excel file has been turned into an image file.")
        except OSError:
            msgbox.showinfo(title="Image Creation Failed", message="Excel file could not be turned into an image file.\nYou can take a screenshot of this page.")
    
    def destroyRequest(self):
        close = msgbox.askyesno(title="Closing Program", message="Are you sure you want to exit the program?")
        if close == True:
            self.quit()
    
    def updateWinScreen(self):
        table.save(xlfile)
        
        sheet.insert_rows(1)
        sheet.insert_rows(1)

        if sheet["C14"].value > sheet["J14"].value:
            pointDiff = int(sheet["C14"].value) - int(sheet["J14"].value)
            winningTeam = homeTeamName
            winText = f"{homeTeamName} wins by {pointDiff}  "
        elif sheet["C14"].value < sheet["J14"].value:
            pointDiff = int(sheet["J14"].value) - int(sheet["C14"].value)
            winningTeam = awayTeamName
            winText = f"{awayTeamName} wins by {pointDiff}  "
        else:
            pointDiff = 0
            winningTeam = "Tied"
            winText = f"Both teams are tied"

        sheet["A1"] = "Winning Team:"
        sheet.merge_cells("B1:C1")
        sheet["B1"] = winningTeam
        sheet["D1"] = "Point Lead:"
        sheet["E1"] = pointDiff

        endTitle.configure(text=winText)

        overviewList = []
        overviewTable = list(sheet.values)
        for row in overviewTable:
            for term in row:
                overviewList.append(term)

        for cell in resultsValues:
            index = resultsValues.index(cell)
            cell.configure(state="normal")
            cell.delete(0, tk.END)
            if str(overviewList[index])!="None":
                cell.insert(1, str(overviewList[index]))
            else:
                cell.insert(1, "")
            cell.configure(state="disabled")
        
        table.save(xlfile)



app = Win()
app.mainloop()